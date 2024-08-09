from flask import Flask, render_template, request, redirect, url_for, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from werkzeug.utils import secure_filename
import os
from openpyxl import load_workbook, Workbook
from sqlalchemy import func
from extensions import db
from models import Question

app = Flask(__name__, static_folder='static', static_url_path='/static')
#app.config['SQLALCHEMY_DATABASE_URI'] = 'postgresql://username:password@hostname/database_name'
app.config['SQLALCHEMY_DATABASE_URI'] = 'postgresql://esg_exam_database_user:Cldtx6hjtBy9El374NXUmdXWUP4k5RVb@dpg-cqgga5qju9rs73cds3fg-a.singapore-postgres.render.com/esg_exam_database'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SECRET_KEY'] = 'ov08fJoEUJM98IDWe42yzmyU0R9jRE35MT5KvbCjJFE'
app.config['SECURITY_PASSWORD_SALT'] = os.environ.get("SECURITY_PASSWORD_SALT", '156025132713563114110595833343437012389')
app.config['UPLOAD_FOLDER'] = 'uploads'

db.init_app(app)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/question_list/<int:page>')
def question_list(page=1):
    try:
        questions = Question.query.paginate(page=page, per_page=20, error_out=False)
    except Exception as e:
        print(f"Error during pagination: {e}")
        questions = None
    return render_template('question_list.html', questions=questions)


@app.route('/edit/<int:id>', methods=['GET', 'POST'])
def edit(id):
    question = Question.query.get_or_404(id)
    if request.method == 'POST':
        question.question_name = request.form['question_name']
        question.question_chapter = request.form['question_chapter']
        question.question_level = request.form['question_level']
        question.question_type = request.form['question_type']
        question.answer = request.form['answer']
        question.option_a = request.form['option_a']
        question.option_b = request.form['option_b']
        question.option_c = request.form['option_c']
        question.option_d = request.form['option_d']
        question.answer_detail = request.form['answer_detail']
        db.session.commit()
        flash('Question updated successfully!')
        return redirect(url_for('question_list', page=1))
    
    # Prepare data for the template
    fields = {
        'question_name': question.question_name,
        'question_chapter': question.question_chapter,
        'question_level': question.question_level,
        'question_type': question.question_type,
        'answer': question.answer,
        'option_a': question.option_a,
        'option_b': question.option_b,
        'option_c': question.option_c,
        'option_d': question.option_d,
        'answer_detail': question.answer_detail,
    }
    return render_template('edit_question.html', fields=fields)

@app.route('/question_delete/<int:id>', methods=['GET', 'POST'])
def question_delete(id):
    question = Question.query.get_or_404(id)
    if request.method == 'POST':
        db.session.delete(question)
        db.session.commit()
        flash('Question deleted successfully!')
        return redirect(url_for('question_list', page=1))
    return render_template('question_delete.html', question=question)

@app.route('/upload', methods=['GET', 'POST'])
def upload():
    if request.method == 'POST':
        file = request.files['file']
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        wb = load_workbook(filepath)
        sheet = wb.active

        max_row = 1
        for row in range(1, sheet.max_row + 1):
            if any(sheet.cell(row=row, column=col).value is not None for col in range(1, 11)):
                max_row = row

        for row in sheet.iter_rows(min_row=2, max_row=max_row, values_only=True):
            question_name = row[0]
            question_chapter = row[1]
            question_level = row[2]
            question_type = row[3]
            answer = row[4]
            option_a = row[5]
            option_b = row[6]
            option_c = row[7]
            option_d = row[8]
            answer_detail = row[9] if len(row) > 9 else None
            
            existing_question = Question.query.filter_by(question_name=question_name).first()
            if existing_question:
                flash(f'This question already exists and was not added: {question_name}', 'error')
                return redirect(url_for('upload'))
            
            question = Question(
                question_name=question_name,
                question_chapter=question_chapter,
                question_level=question_level,
                question_type=question_type,
                answer=answer,
                option_a=option_a,
                option_b=option_b,
                option_c=option_c,
                option_d=option_d,
                answer_detail=answer_detail,
            )
            db.session.add(question)

        db.session.commit()
        flash('Questions uploaded successfully!')
        return redirect(url_for('question_list', page=1))
    
    return render_template('upload.html')

@app.route('/download')
def download():
    questions = Question.query.all()
    wb = Workbook()
    ws = wb.active
    ws.append(['題目', '題目對應章節', '難度', '題型', '答案', '選項A', '選項B', '選項C', '選項D', '詳解'])
    for question in questions:
        ws.append([question.question_name, question.question_chapter, question.question_level, question.question_type, question.answer, question.option_a, question.option_b, question.option_c, question.option_d, question.answer_detail])
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], 'questions_backup.xlsx')
    wb.save(filepath)
    return send_file(filepath, as_attachment=True)

@app.route('/quiz_generation', methods=['GET', 'POST'])
def quiz_generation():
    if request.method == 'POST':
        selected_chapters = request.form.getlist('chapter')
        selected_levels = request.form.getlist('level')
        selected_types = request.form.getlist('type')
        num_questions = int(request.form['num_questions'])

        # 查询数据库以获取符合条件的问题
        questions = Question.query.filter(
            Question.question_chapter.in_(selected_chapters),
            Question.question_level.in_(selected_levels),
            Question.question_type.in_(selected_types)
        ).order_by(func.random()).limit(num_questions).all()

        return render_template('generated_quiz.html', questions=questions)
    
    chapters = db.session.query(Question.question_chapter).distinct().all()
    levels = db.session.query(Question.question_level).distinct().all()
    types = db.session.query(Question.question_type).distinct().all()

    return render_template('quiz_generation.html', chapters=chapters, levels=levels, types=types)



if __name__ == '__main__':
    with app.app_context():
        db.create_all()  # 建立資料表
    app.run(debug=True)

